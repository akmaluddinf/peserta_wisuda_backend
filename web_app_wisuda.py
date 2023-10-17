import openpyxl
from flask import Flask, render_template, send_from_directory, request, jsonify, json
from fpdf import FPDF
import pandas as pd
from flask_cors import CORS
import locale
from datetime import datetime
import logging
import gspread
from oauth2client.service_account import ServiceAccountCredentials

app = Flask(__name__)

CORS(app, resources={r"/search": {"origins": "http://localhost:8002"}})  # Sesuaikan dengan URL React Anda

EXCEL_PATH = "wisudafix.xlsx"
PDF_PATH = "static/result.pdf"
LOG_FILE_PATH = "download_log.txt"  # Menambahkan path untuk file log

# Konfigurasi logging
logging.basicConfig(filename=LOG_FILE_PATH, level=logging.INFO, format='%(asctime)s - %(message)s')

# Path ke file kredensial JSON yang telah diunduh
credentials_path = 'minime-pro-01-f29431f9e490.json'

# Otorisasi menggunakan kredensial
scope = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
credentials = ServiceAccountCredentials.from_json_keyfile_name(credentials_path, scope)
gc = gspread.authorize(credentials)


# ID Google Sheet (dapat ditemukan di URL)
sheet_id = '1_yQE7ys1_Jx3O0VmgHYzaLxHi-MwuUgIklgjmS0jjic'

# Nama worksheet di Google Sheet
worksheet_name = 'Master Wisuda 2324 G1'  # Ganti dengan nama worksheet Anda

# Fungsi untuk membaca data dari Google Sheet
def read_google_sheet():
    try:
        # Buka Google Sheet
        worksheet = gc.open_by_key(sheet_id).worksheet(worksheet_name)

        # Ambil semua nilai dari worksheet
        values = worksheet.get_all_values()

        # Header kolom dari baris pertama
        header = values[0]
        # print("Header kolom:", header)

        # Buat DataFrame dari data
        df = pd.DataFrame(values[1:], columns=header)

        # Cetak beberapa baris data pertama
        # print("Data pertama:")
        # print(df.head())

        return df

    except Exception as e:
        logging.error(f"Error reading Google Sheet: {e}")
        return None

# Fungsi untuk mencari data dari Google Sheet
def search_google_sheet(keyword):
    results = []

    # Baca data dari Google Sheet
    df = read_google_sheet()

    if df is not None:
        # Lakukan pencarian di DataFrame
        filtered_df = df.loc[df['NIM'].str.lower() == keyword.lower()]

        # Konversi hasil pencarian menjadi daftar baris
        results = filtered_df.to_dict(orient='records')

        # Log hasil pencarian
        logging.info(f"Search for NIM {keyword} - Entries found: {len(results)}")

    return results

@app.route('/', methods=['GET', 'POST'])
def index():
    message = ""
    found = False

    if request.method == 'POST':
        keyword = request.form.get('keyword')
        results = search_excel(keyword)

        if results:
             # Menghapus duplikat dari hasil pencarian
            unique_results = []
            seen_nims = set()
            for row in results:
                nim = row[1]
                if nim not in seen_nims:
                    unique_results.append(row)
                    seen_nims.add(nim)

            generate_pdf(keyword, unique_results, found)
            found = True
        else:
            message = f"No entries found for the keyword: '{keyword}'."
            logging.warning(f"No entries found for the keyword: '{keyword}'.")

    return render_template('index.html', found=found, message=message)

# Tambahkan rute untuk mengatasi permintaan OPTIONS
@app.route('/search', methods=['GET', 'OPTIONS'])
def search():
    if request.method == 'OPTIONS':
        response = app.response_class(
            response=json.dumps({'message': 'CORS preflight OK'}),
            status=200,
            mimetype='application/json'
        )
        return response
    elif request.method == 'GET':
        keyword = request.args.get('nim')
        # results = search_google_sheet(keyword) #Jika get data dari google sheet
        results = search_excel(keyword) #Jika get data dari file excel
        found = False  # Inisialisasi variabel found di sini

        if results:
            # Menghapus duplikat dari hasil pencarian
            unique_results = []
            seen_nims = set()
            for row in results:
                # nim = row['NIM']  # jika get data dari google sheet
                nim = row[1]  # jika get data dari file excel
                if nim not in seen_nims:
                    unique_results.append({
                        "NIM": row[1],
                        "Nama": row[2],
                        "Fakultas": row[4],
                        "Program Studi": row[3],
                        "Ukuran Almamater": row[9],
                        "Nomor Urut": row[11],
                        "Mengisi Tracer Study": row[12],
                        "Status Tagihan Wisuda": row[10],
                        "Waktu Bayar": row[14],
                        "Sesi Wisuda": row[16]

                        # Jika get data dari google sheet
                        
                        # "NIM": row['NIM'],
                        # "Nama": row['NAMA MAHASISWA'],
                        # "Fakultas": row['FAKULTAS'],
                        # "Program Studi": row['PROGRAM STUDI'],
                        # "Ukuran Almamater": row['UK. ALMAMATER'],
                        # "Nomor Urut": row['NO. Kursi'],
                        # "Mengisi Tracer Study": row['Tracer Study'],
                        # "Status Tagihan Wisuda": row['STATUS TAGIHAN WISUDA'],
                        # "Waktu Bayar": row['WAKTU BAYAR']
                    })
                    seen_nims.add(nim)

            generate_pdf(keyword, unique_results, found)
            found = True
        else:
            found = False
            logging.warning(f"No entries found for the keyword: '{keyword}'.")
            return jsonify({"message": f"No entries found for the keyword: '{keyword}'."})

        return jsonify({"found": found, "mahasiswa": unique_results})


@app.route('/download')
def download():
    keyword = request.args.get('nim')
    status_code = 200  # Tetapkan status code default

    try:
        # Coba untuk mendownload file
        pdf_filename = f"BUKTI_WISUDA_{keyword}.pdf"
        return send_from_directory('static', 'result.pdf', as_attachment=True, download_name=pdf_filename)
    except Exception as e:
        status_code = 500  # Perbarui status code untuk error
        logging.error(f"Error during file download: {e}")

    # Log informasi download
    log_info = {"NIM": keyword, "Status Code": status_code, "Timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S")}
    logging.info(json.dumps(log_info))

    # Kembalikan respons yang sesuai
    if status_code == 200:
        return jsonify({"message": "Download berhasil"})
    else:
        return jsonify({"message": "Error selama proses download"}), 500

def search_excel(keyword):
    results = []
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb.active

    for row in ws.iter_rows(values_only=True):
        if str(keyword).lower() == str(row[1]).lower():
            results.append(row)
            logging.info(f"Search for NIM {keyword} - Entry found: {row}")
    return results



def generate_pdf(keyword, results, found):
    pdf = FPDF()
    pdf.add_page()
    pdf.set_auto_page_break(auto=False)
    pdf.set_font("Times", size=12)


    # Tambahkan gambar kop surat sebagai header
    pdf.image("header.png", x=10, y=5, w=190)

    #Menambahkan space beetween
    pdf.cell(200, 35, ln=True)

    # Ukuran dan posisi gambar logo di sebelah kiri atas
    left_logo_x = 10
    left_logo_y = 10
    left_logo_w = 20

    # Ukuran dan posisi gambar logo di sebelah kanan atas
    right_logo_w = 20
    right_logo_x = pdf.w - 10 - right_logo_w  # Letakkan logo di kanan atas
    right_logo_y = 10

    # Tinggi sel kosong antara paragraf
    space_between_paragraphs = -3 

    # Mengatur font size dan membuat tulisan bold
    pdf.set_font("Arial", size=12, style='B')  # Font size 14 dan bold

    # Menambahkan judul di tengah-tengah
    pdf.cell(200, 10, txt="PESERTA WISUDA SARJANA (S1) DAN PASCASARJANA (S2 & S3)", ln=True, align='C')
    pdf.cell(200, space_between_paragraphs, ln=True)

    #Mengurangi space beetween
    pdf.cell(200, -1, ln=True)
    pdf.cell(200, 10, txt="UNIVERSITAS PASUNDAN GELOMBANG I TAHUN AKADEMIK 2023/2024", ln=True, align='C')
    pdf.cell(200, space_between_paragraphs, ln=True)

    # Mengembalikan ke font reguler
    pdf.set_font("Arial", size=8)
    pdf.cell(200, 10, txt="Sekretariat: Jl. Tamansari No. 4-8 Bandung, Call Center: 0811960193, Email: rektorat@unpas.ac.id", ln=True, align='C')
    pdf.cell(200, -7, ln=True)
    pdf.cell(200, 10, txt="Email: rektorat@unpas.ac.id Website: www.unpas.ac.id", ln=True, align='C')

    # Menambahkan Teks diatas
    pdf.cell(200, 5, ln=True)
    # Mengembalikan ke font reguler
    pdf.set_font("Arial", size=11)
    # Menambahkan indent
    indent = 15

    pdf.set_x(pdf.get_x() + indent)  # Geser kursor ke posisi indent
    #Menambahkan teks keterangan
    pdf.cell(200, 10, txt="Selamat Anda telah terdaftar sebagai Peserta Wisuda Universitas Pasundan Gelombang I ", ln=True)
    #Menambahkan space beetween
    pdf.cell(200, -4, ln=True)
    pdf.set_x(pdf.get_x() + indent)  # Geser kursor ke posisi indent
    pdf.cell(200, 10, txt="Tahun Akademik 2023/2024, dengan data sebagai berikut:", ln=True)

    #Menambahkan space beetween
    pdf.cell(200, 0, ln=True)
    # Mengatur font size dan membuat tulisan bold
    pdf.set_font("Arial", size=12, style='B')
    # Menambahkan "DATA MAHASISWA" di tengah
    pdf.cell(200, 10, txt="DATA WISUDAWAN/WISUDAWATI", ln=True, align='C')
    pdf.cell(200, 0, ln=True)

    # Mengembalikan ke font reguler
    pdf.set_font("Arial", size=11)


    # Iterasi melalui hasil pencarian
    for row in results:

        # Mengatur lebar kolom untuk NIM
        col_width = 50

        #Menambahkan indent
        indent = 15

        pdf.set_x(pdf.get_x() + indent)  # Geser kursor ke posisi indent
        # Mengisi kolom NIM
        pdf.cell(col_width, 10, txt="NIM", align='L')
        pdf.cell(200 - col_width, 10, txt=": " + str(row['NIM']), ln=True)
        #Mengurangi space beetween
        pdf.cell(200, -3, ln=True)

        pdf.set_x(pdf.get_x() + indent)  # Geser kursor ke posisi indent
        # Mengisi kolom Nama dengan lebar kolom yang sama
        pdf.cell(col_width, 10, txt="Nama", align='L')
        pdf.cell(200 - col_width, 10, txt=": " + str(row['Nama']), ln=True)
        #Mengurangi space beetween  
        pdf.cell(200, -3, ln=True)

        pdf.set_x(pdf.get_x() + indent)  # Geser kursor ke posisi indent
        # Mengisi kolom Program Studi dengan lebar kolom yang sama
        pdf.cell(col_width, 10, txt="Program Studi", align='L')
        pdf.cell(200 - col_width, 10, txt=": " + str(row['Program Studi']), ln=True)
        #Mengurangi space beetween
        pdf.cell(200, -3, ln=True)

        pdf.set_x(pdf.get_x() + indent)  # Geser kursor ke posisi indent
        # Mengisi kolom Fakultas dengan lebar kolom yang sama
        pdf.cell(col_width, 10, txt="Fakultas", align='L')
        pdf.cell(200 - col_width, 10, txt=": " + str(row['Fakultas']), ln=True)
        #Mengurangi space beetween
        pdf.cell(200, -3, ln=True)

        pdf.set_x(pdf.get_x() + indent)  # Geser kursor ke posisi indent
        # Mengisi kolom Size Almamater dengan lebar kolom yang sama
        pdf.cell(col_width, 10, txt="Ukuran Toga", align='L')
        pdf.cell(200 - col_width, 10, txt=": " + str(row['Ukuran Almamater']), ln=True)
        #Mengurangi space beetween
        pdf.cell(200, -3, ln=True)

        pdf.set_x(pdf.get_x() + indent)  # Geser kursor ke posisi indent
        # Mengisi kolom Nomor Urut dengan lebar kolom yang sama
        pdf.cell(col_width, 10, txt="Nomor Urut/Kursi", align='L')
        pdf.cell(200 - col_width, 10, txt=": " + str(row['Nomor Urut']), ln=True)
        #Mengurangi space beetween
        pdf.cell(200, -3, ln=True)

        pdf.set_x(pdf.get_x() + indent)  # Geser kursor ke posisi indent
        # Mengisi kolom Sesi Wisuda dengan lebar kolom yang sama
        pdf.cell(col_width, 10, txt="Sesi Wisuda", align='L')
        pdf.cell(200 - col_width, 10, txt=": " + "Sesi 1", ln=True)
        #Mengurangi space beetween
        pdf.cell(200, -3, ln=True)

        pdf.set_x(pdf.get_x() + indent)  # Geser kursor ke posisi indent
        # Mengisi kolom Lokasi Wisuda dengan lebar kolom yang sama
        pdf.cell(col_width, 10, txt="Lokasi Wisuda", align='L')
        pdf.cell(200 - col_width, 10, txt=": " + "Sasana Budaya Ganesha (SABUGA)", ln=True)
        #Mengurangi space beetween
        pdf.cell(200, -3, ln=True)

        pdf.set_x(pdf.get_x() + indent)  # Geser kursor ke posisi indent
        # Mengisi kolom Lokasi Wisuda dengan lebar kolom yang sama
        pdf.cell(col_width, 10, txt="Waktu Pelaksanaan", align='L')

        sesi_wisuda = str(row.get('Sesi Wisuda', ''))
        # Menentukan waktu pelaksanaan berdasarkan sesi
        if sesi_wisuda == "Sesi 1":
            waktu_pelaksanaan = "Sabtu, 11 November 2023, 08.00 - 11.00"
        else:
            waktu_pelaksanaan = "Sabtu, 11 November 2023, 14.00 s.d. Selesai"

        pdf.cell(200 - col_width, 10, txt=": " + waktu_pelaksanaan, ln=True)
        #Mengurangi space beetween
        pdf.cell(200, -3, ln=True)

        pdf.set_x(pdf.get_x() + indent)  # Geser kursor ke posisi indent
        # Mengisi kolom Lokasi Wisuda dengan lebar kolom yang sama
        pdf.cell(col_width, 10, txt="Status Tagihan Wisuda", align='L')
        pdf.cell(200 - col_width, 10, txt=": " + str(row['Status Tagihan Wisuda']), ln=True)
        #Mengurangi space beetween
        pdf.cell(200, -3, ln=True)

        pdf.set_x(pdf.get_x() + indent)  # Geser kursor ke posisi indent
        # Mengisi kolom Waktu Wisuda dengan lebar kolom yang sama
        pdf.cell(col_width, 10, txt="Tanggal Bayar", align='L')

        # Ambil nilai 'Waktu Bayar' dari row
        waktu_bayar = str(row.get('Waktu Bayar', ''))  # Mendapatkan nilai, default ke string kosong jika None

        # print('waktu bayar:', waktu_bayar)

        if waktu_bayar is not None and waktu_bayar.strip():  # Jika waktu_bayar tidak kosong
            # Konversi string tanggal ke objek datetime
            try:
                # Ubah format tanggal sesuai dengan format yang diberikan dari Google Sheets
                # (SET DI GOOGLE SHEETS DENGAN FORMAT: (BULAN/TANGGAL/TAHUN JAM:MENIT:DETIK) CONTOH: 10/03/2023 14:57:34
                # tanggal_datetime = datetime.strptime(str(waktu_bayar), "%m/%d/%Y %H:%M:%S") #Jika get dari data google sheets

                # Ubah format tanggal sesuai dengan format yang diberikan dari file excel
                #Format waktu di excel: (TAHUN-BULAN-TANGGAL JAM:MENIT:DETIK) CONTOH: 2023-10-03 14:57:34
                tanggal_datetime = datetime.strptime(str(waktu_bayar), "%Y-%m-%d %H:%M:%S") #Jika get dari data file excel

                # Set lokal ke Bahasa Indonesia
                locale.setlocale(locale.LC_TIME, 'id_ID')

                # Format tanggal dalam Bahasa Indonesia
                tanggal_format_indonesia = tanggal_datetime.strftime("%d %B %Y %H:%M")

                # tampilkan tanggal
                text_to_display = f": {tanggal_format_indonesia}"

                # Tampilkan di PDF
                pdf.cell(200 - col_width, 10, txt=text_to_display, ln=True)
            except ValueError as e:
                # Tampilkan pesan bahwa format tanggal tidak valid
                pdf.cell(200 - col_width, 10, txt=": BELUM LUNAS", ln=True)
        else:
            # Tampilkan pesan bahwa waktu bayar tidak tersedia
            pdf.cell(200 - col_width, 10, txt=": BELUM LUNAS", ln=True)


        #Mengurangi space beetween
        pdf.cell(200, -3, ln=True)

        pdf.set_x(pdf.get_x() + indent)  # Geser kursor ke posisi indent

        pdf.cell(col_width, 10, txt="Mengisi Tracer Study", align='L')

        # Ambil nilai 'Mengisi Tracer Study' dari row
        mengisi_tracer_study = str(row.get('Mengisi Tracer Study', ''))

        # Periksa apakah nilai '#N/A', jika ya, ganti dengan 'Belum Mengisi'
        if mengisi_tracer_study != "Ya, Sudah Mengisi":
            mengisi_tracer_study = "Belum Mengisi"

        # Tampilkan nilai pada PDF
        pdf.cell(200 - col_width, 10, txt=": " + mengisi_tracer_study, ln=True)



    #Menambahkan space beetween
    pdf.cell(200, 0, ln=True)
        
    pdf.set_x(pdf.get_x() + indent)  # Geser kursor ke posisi indent
    #Menambahkan teks keterangan
    pdf.cell(200, 10, txt="Surat Keterangan ini bisa digunakan sebagai bukti untuk pengambilan perlengkapan Peserta", ln=True)

    #Menambahkan space beetween
    pdf.cell(200, -4, ln=True)
    pdf.set_x(pdf.get_x() + indent)  # Geser kursor ke posisi indent
    pdf.cell(200, 10, txt="Wisuda Universitas Pasundan Gelombang I Tahun Akademik 2023/2024.", ln=True)

    #Menambahkan space beetween
    pdf.cell(200, 0, ln=True)

    #Menambahkan space beetween
    pdf.cell(200, 0, ln=True)
    pdf.set_x(pdf.get_x() + indent)  # Geser kursor ke posisi indent
    pdf.cell(200, 10, txt="Ceklis pengambilan perlengkapan wisuda:", ln=True)

    #Menambahkan space beetween
    pdf.cell(200, -3, ln=True)
    pdf.set_x(pdf.get_x() + indent)  # Geser kursor ke posisi indent
    pdf.cell(200, 10, txt="[  ] Toga", ln=True)

    #Menambahkan space beetween
    pdf.cell(200, -3, ln=True)
    pdf.set_x(pdf.get_x() + indent)  # Geser kursor ke posisi indent
    pdf.cell(200, 10, txt="[  ] Pin", ln=True)

    #Menambahkan space beetween
    pdf.cell(200, -3, ln=True)
    pdf.set_x(pdf.get_x() + indent)  # Geser kursor ke posisi indent
    pdf.cell(200, 10, txt="[  ] Undangan Wisuda", ln=True)



    # Menggeser kursor ke posisi kanan bawah
    pdf.set_xy(130, 225)  # Sesuaikan posisi dan koordinatnya

    # Set lokal ke Bahasa Indonesia
    locale.setlocale(locale.LC_TIME, 'id_ID')

    # Menambahkan tanggal download
    tanggal_download = datetime.now().strftime("%d %B %Y")
    
    # Menambahkan tanggal
    pdf.cell(60, 40, txt=f"Bandung, {tanggal_download}", ln=True, align='R')

    # Menambahkan panitia
    pdf.cell(160, 10, txt="Panitia", ln=True, align='R')

    # geser kursor ke posisi bawah
    pdf.set_y(-25)  # Sesuaikan nilai -40 sesuai dengan tinggi gambar footer

    # Tambahkan gambar footer
    pdf.image("footer.png", x=10, w=190)


    pdf.output(PDF_PATH)

if __name__ == "__main__":
    app.run(debug=True, port=8001)
